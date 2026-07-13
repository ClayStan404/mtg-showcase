#!/usr/bin/env python3
"""解析 WPS 库存协作 Excel → inventory/*.txt（供 build_data 使用）。

约定（与 templates/WPS库存协作模板.xlsx 一致）：
  - 跳过工作表：说明、模板、名称以「模板」开头
  - 昵称/城市/联系：A4/A5/A6 标签，B 列值（或前几行「昵称」键值）
  - 卡表：表头行含「系列」「编号」，其下为数据
  - 语言：e=英 z=中 j=日 o=其他（空默认 e）
  - 闪：空/0=否 1=是
  - 数量：空=1
"""

from __future__ import annotations

import argparse
import re
import sys
from collections import OrderedDict
from pathlib import Path
from typing import Any

# 允许直接 python scripts/parse_wps_excel.py
sys.path.insert(0, str(Path(__file__).resolve().parent))

from inventory_format import (  # noqa: E402
    LANG_TOKEN,
    ParseError,
    cell_str,
    normalize_foil,
    normalize_lang,
    normalize_qty,
    slugify,
    validate_meta,
)

try:
    from openpyxl import load_workbook
except ImportError:
    print("需要 openpyxl：pip install openpyxl", file=sys.stderr)
    sys.exit(1)

ROOT = Path(__file__).resolve().parents[1]
SKIP_SHEET_NAMES = {"说明", "使用说明", "填写说明", "对照表（一般不用改）"}
SKIP_SHEET_PREFIXES = ("模板", "template", "_")


def should_skip_sheet(name: str) -> bool:
    n = (name or "").strip()
    if not n or n in SKIP_SHEET_NAMES:
        return True
    low = n.lower()
    for p in SKIP_SHEET_PREFIXES:
        if n.startswith(p) or low.startswith(p):
            return True
    return False


def find_meta(ws) -> dict[str, str]:
    """从工作表前 15 行找 昵称/城市/联系。"""
    meta = {"seller": "", "city": "", "contact": ""}
    key_map = {
        "昵称": "seller",
        "卖家昵称": "seller",
        "卖家": "seller",
        "seller": "seller",
        "城市": "city",
        "city": "city",
        "联系": "contact",
        "联系方式": "contact",
        "contact": "contact",
        "微信": "contact",
    }
    for row in ws.iter_rows(min_row=1, max_row=15, max_col=4, values_only=True):
        if not row:
            continue
        label = cell_str(row[0])
        val = cell_str(row[1]) if len(row) > 1 else ""
        if label in key_map and val:
            meta[key_map[label]] = val
    return meta


def find_header_row(ws) -> tuple[int, dict[str, int]] | None:
    """返回 (1-based row, {field: col_index 0-based})。"""
    aliases = {
        "系列": "set",
        "系列缩写": "set",
        "set": "set",
        "编号": "number",
        "收集编号": "number",
        "number": "number",
        "语言": "lang",
        "lang": "lang",
        "闪": "foil",
        "是否闪卡": "foil",
        "闪卡": "foil",
        "foil": "foil",
        "数量": "qty",
        "qty": "qty",
        "数量*": "qty",
        "备注": "note",
        "备注（可选）": "note",
        "note": "note",
    }
    for r_idx, row in enumerate(ws.iter_rows(min_row=1, max_row=30, max_col=12, values_only=True), 1):
        if not row:
            continue
        colmap: dict[str, int] = {}
        for c_idx, cell in enumerate(row):
            key = aliases.get(cell_str(cell))
            if key:
                colmap[key] = c_idx
        if "set" in colmap and "number" in colmap:
            return r_idx, colmap
    return None


def parse_sheet(ws, sheet_name: str) -> tuple[dict[str, str], list[dict[str, Any]], list[str]]:
    """返回 (meta, cards, errors)。"""
    errors: list[str] = []
    meta = find_meta(ws)
    header = find_header_row(ws)
    if not header:
        errors.append(f"[{sheet_name}] 未找到表头（需要「系列」「编号」列）")
        return meta, [], errors

    header_row, colmap = header
    cards: list[dict[str, Any]] = []

    for r_idx, row in enumerate(
        ws.iter_rows(min_row=header_row + 1, max_col=12, values_only=True),
        header_row + 1,
    ):
        if not row or all(cell_str(c) == "" for c in row):
            continue
        try:
            set_code = cell_str(row[colmap["set"]]).lower()
            number = cell_str(row[colmap["number"]])
            if not set_code and not number:
                continue
            if not set_code or not number:
                raise ParseError("系列与编号须同时填写", r_idx)

            lang_raw = row[colmap["lang"]] if "lang" in colmap else ""
            foil_raw = row[colmap["foil"]] if "foil" in colmap else ""
            qty_raw = row[colmap["qty"]] if "qty" in colmap else ""

            lang = normalize_lang(lang_raw)
            foil = normalize_foil(foil_raw)
            qty = normalize_qty(qty_raw)

            cards.append(
                {
                    "set": set_code,
                    "number": number,
                    "lang": lang,
                    "foil": foil,
                    "quantity": qty,
                    "line": r_idx,
                }
            )
        except ParseError as e:
            errors.append(f"[{sheet_name}] 第{r_idx}行：{e}")

    if not cards and not errors:
        # 空库存允许
        pass
    errors.extend(validate_meta(meta, sheet_name))
    return meta, cards, errors


def merge_cards(cards: list[dict[str, Any]]) -> list[dict[str, Any]]:
    merged: OrderedDict[str, dict[str, Any]] = OrderedDict()
    for c in cards:
        key = f"{c['set']}|{c['number']}|{c['lang']}|{'f' if c['foil'] else 'nf'}"
        if key in merged:
            merged[key]["quantity"] += c["quantity"]
        else:
            merged[key] = dict(c)
    return list(merged.values())


def cards_to_txt(meta: dict[str, str], cards: list[dict[str, Any]]) -> str:
    lines = [
        f"# seller: {meta.get('seller') or ''}",
        f"# city: {meta.get('city') or ''}",
        f"# contact: {meta.get('contact') or ''}",
        "#",
        "# 由 parse_wps_excel.py 生成 — 语言 e/z/j/o（空默认e） 闪 0/1  数量默认1",
        "",
    ]
    for c in cards:
        # 内部 txt：与 build_data 兼容
        # lang: en→e, zhs→z, ja→j, other→o
        lang_token = LANG_TOKEN.get(c["lang"], "e")
        foil_token = "1" if c["foil"] else "0"
        qty = c["quantity"]
        prefix = f"{qty}x " if qty != 1 else ""
        line = f"{prefix}{c['set']} {c['number']} {lang_token} {foil_token}"
        lines.append(line)
    lines.append("")
    return "\n".join(lines)


def parse_workbook(path: Path) -> tuple[dict[str, tuple[dict, list]], list[str]]:
    """返回 {sheet_name: (meta, cards)}, errors"""
    wb = load_workbook(path, read_only=True, data_only=True)
    result: dict[str, tuple[dict, list]] = {}
    errors: list[str] = []
    try:
        for name in wb.sheetnames:
            if should_skip_sheet(name):
                continue
            ws = wb[name]
            meta, cards, errs = parse_sheet(ws, name)
            errors.extend(errs)
            if errs:
                continue
            result[name] = (meta, merge_cards(cards))
    finally:
        wb.close()
    return result, errors


def main() -> int:
    parser = argparse.ArgumentParser(description="WPS Excel → inventory/*.txt")
    parser.add_argument("xlsx", type=Path, help="WPS 导出的 xlsx 路径")
    parser.add_argument(
        "-o",
        "--out-dir",
        type=Path,
        default=ROOT / "inventory",
        help="输出 inventory 目录（默认 ./inventory）",
    )
    parser.add_argument(
        "--dry-run",
        action="store_true",
        help="只校验并打印摘要，不写文件",
    )
    args = parser.parse_args()

    if not args.xlsx.is_file():
        print(f"文件不存在: {args.xlsx}", file=sys.stderr)
        return 1

    print(f"解析 {args.xlsx} …")
    sheets, errors = parse_workbook(args.xlsx)

    if errors:
        print(f"\n❌ 校验失败（{len(errors)} 个问题）：", file=sys.stderr)
        for e in errors:
            print(f"  · {e}", file=sys.stderr)
        return 1

    if not sheets:
        print("没有可导入的工作表（是否只有「说明/模板」？）", file=sys.stderr)
        return 1

    print("\n✅ 校验通过：")
    for name, (meta, cards) in sheets.items():
        total = sum(c["quantity"] for c in cards)
        print(
            f"  · {name}: {meta.get('seller')} / {meta.get('city')} "
            f"— {len(cards)} 种 / {total} 张"
        )

    if args.dry_run:
        print("\n(--dry-run，未写入文件)")
        return 0

    args.out_dir.mkdir(parents=True, exist_ok=True)
    # 仅写入本次 xlsx 中的卖家；不删除其他 txt（避免误删）
    used_fnames: set[str] = set()
    for name, (meta, cards) in sheets.items():
        seller = meta.get("seller") or name
        fname = slugify(seller, slugify(name, "seller")) + ".txt"
        # 工作表名若是英文 id 优先用表名
        if re.fullmatch(r"[A-Za-z0-9_-]+", name.strip()):
            fname = name.strip() + ".txt"
        # 冲突检测：同名卖家 slugify 后相同会静默覆盖丢数据，加后缀避免
        if fname in used_fnames:
            stem, _, ext = fname.rpartition(".")
            i = 2
            while f"{stem}_{i}.{ext}" in used_fnames:
                i += 1
            fname = f"{stem}_{i}.{ext}"
            print(f"  ⚠ 工作表「{name}」文件名冲突，改用 {fname}", file=sys.stderr)
        used_fnames.add(fname)
        path = args.out_dir / fname
        path.write_text(cards_to_txt(meta, cards), encoding="utf-8")
        print(f"  写入 {path.relative_to(ROOT) if path.is_relative_to(ROOT) else path}")

    print("\n下一步: python3 scripts/build_data.py")
    return 0


if __name__ == "__main__":
    sys.exit(main())
