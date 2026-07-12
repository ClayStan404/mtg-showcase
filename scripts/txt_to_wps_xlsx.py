#!/usr/bin/env python3
"""把「Excel 列顺序」的纯文本库存 txt → WPS 协作模板格式 xlsx。

读取格式（每行）: 系列 编号 语言 闪 数量
输出: 与 templates/WPS库存协作模板.xlsx 结构一致的 xlsx，可直接导入 WPS。

用法:
  python3 scripts/txt_to_wps_xlsx.py claystan.txt              # 输出 claystan.xlsx
  python3 scripts/txt_to_wps_xlsx.py claystan.txt -o 库存.xlsx
"""

from __future__ import annotations

import argparse
import sys
from pathlib import Path

sys.path.insert(0, str(Path(__file__).resolve().parent))

from parse_excel_order_txt import (  # noqa: E402
    merge_cards,
    parse_file,
)
from inventory_format import LANG_TOKEN, validate_meta  # noqa: E402

try:
    import openpyxl
    from openpyxl.styles import Font
except ImportError:
    print("需要 openpyxl：pip install openpyxl", file=sys.stderr)
    sys.exit(1)

ROOT = Path(__file__).resolve().parents[1]
FONT = "微软雅黑"


def build_sheet(ws, meta: dict[str, str], cards: list[dict]) -> None:
    """在 ws 上写入 WPS 模板格式的库存数据。"""
    # Row 1: title (merged A1:F1)
    ws["A1"] = "MTG 库存"
    ws["A1"].font = Font(name=FONT, size=14, bold=True)
    ws.merge_cells("A1:F1")

    # Row 2: instructions (merged A2:F2)
    ws["A2"] = "语言 e/z/j/o（空=e）｜闪 0/1（空=0）｜数量空=1｜卖光删行｜无权限 QQ 417592443"
    ws["A2"].font = Font(name=FONT, size=10)
    ws.merge_cells("A2:F2")

    # Row 4-6: metadata (label bold, value merged B:C)
    labels = [
        (4, "昵称", meta.get("seller", "")),
        (5, "城市", meta.get("city", "")),
        (6, "联系", meta.get("contact", "")),
    ]
    for row, label, value in labels:
        ws.cell(row=row, column=1, value=label).font = Font(name=FONT, size=11, bold=True)
        ws.cell(row=row, column=2, value=value).font = Font(name=FONT, size=11)
        ws.merge_cells(start_row=row, start_column=2, end_row=row, end_column=3)

    # Row 8: headers
    headers = ["系列", "编号", "语言", "闪", "数量", "备注"]
    for col, header in enumerate(headers, 1):
        ws.cell(row=8, column=col, value=header).font = Font(name=FONT, size=11, bold=True)

    # Row 9+: data
    lang_token = LANG_TOKEN
    for i, c in enumerate(cards):
        row = 9 + i
        ws.cell(row=row, column=1, value=c["set"]).font = Font(name=FONT, size=11)
        ws.cell(row=row, column=2, value=c["number"]).font = Font(name=FONT, size=11)
        ws.cell(row=row, column=3, value=lang_token.get(c["lang"], "e")).font = Font(name=FONT, size=11)
        ws.cell(row=row, column=4, value=1 if c["foil"] else 0).font = Font(name=FONT, size=11)
        ws.cell(row=row, column=5, value=c["quantity"]).font = Font(name=FONT, size=11)

    # Column widths
    ws.column_dimensions["A"].width = 12
    ws.column_dimensions["B"].width = 14
    ws.column_dimensions["C"].width = 8
    ws.column_dimensions["D"].width = 6
    ws.column_dimensions["E"].width = 8
    ws.column_dimensions["F"].width = 20


def main() -> int:
    parser = argparse.ArgumentParser(description="txt → WPS 格式 xlsx")
    parser.add_argument("txt", type=Path, help="输入 txt（系列 编号 语言 闪 数量）")
    parser.add_argument("-o", "--output", type=Path, default=None, help="输出 xlsx 路径")
    parser.add_argument("--sheet-name", default="", help="工作表名（默认用 seller）")
    args = parser.parse_args()

    if not args.txt.is_file():
        print(f"文件不存在: {args.txt}", file=sys.stderr)
        return 1

    print(f"解析 {args.txt} …")
    meta, cards, errors = parse_file(args.txt)
    errors.extend(validate_meta(meta, args.txt.name))

    if errors:
        print(f"\n❌ 校验失败（{len(errors)} 个问题）：", file=sys.stderr)
        for e in errors:
            print(f"  · {e}", file=sys.stderr)
        return 1

    cards = merge_cards(cards)
    total = sum(c["quantity"] for c in cards)
    print(f"✅ {meta.get('seller')} / {meta.get('city')} — {len(cards)} 种 / {total} 张")

    sheet_name = args.sheet_name or meta.get("seller") or args.txt.stem
    output = args.output or ROOT / f"{args.txt.stem}.xlsx"

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = sheet_name[:31]  # Excel sheet name max 31 chars
    build_sheet(ws, meta, cards)
    wb.save(output)
    print(f"已写入 {output}")
    return 0


if __name__ == "__main__":
    sys.exit(main())
