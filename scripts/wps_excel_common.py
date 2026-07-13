#!/usr/bin/env python3
"""parse_wps_excel / parse_wps_wants_excel 共享的 WPS Excel 解析逻辑。

提取公共的跳过规则、meta/header 查找、workbook 遍历、文件写入（含冲突检测），
两个脚本各自只定义差异：key_map（seller vs buyer）、header aliases（求购多 must 列）、
parse_sheet 的字段、merge/to_txt。
"""

from __future__ import annotations

import re
import sys
from pathlib import Path

try:
    from openpyxl import load_workbook
except ImportError:
    print("需要 openpyxl：pip install openpyxl", file=sys.stderr)
    sys.exit(1)

sys.path.insert(0, str(Path(__file__).resolve().parent))
from inventory_format import cell_str, slugify  # noqa: E402

SKIP_SHEET_NAMES = {"说明", "使用说明", "填写说明", "对照表（一般不用改）"}
SKIP_SHEET_PREFIXES = ("模板", "template", "_")


def should_skip_sheet(name: str) -> bool:
    n = (name or "").strip()
    if not n or n in SKIP_SHEET_NAMES:
        return True
    low = n.lower()
    for p in SKIP_SHEET_PREFIXES:
        if n.startswith(p) or low.startswith(p):
            return True
    return False


def find_meta(ws, key_map: dict[str, str], meta_keys: tuple[str, ...]) -> dict[str, str]:
    """从工作表前 15 行找 meta。key_map: {单元格标签: meta_key}；meta_keys: 初始 meta 的 keys。"""
    meta = {k: "" for k in meta_keys}
    for row in ws.iter_rows(min_row=1, max_row=15, max_col=4, values_only=True):
        if not row:
            continue
        label = cell_str(row[0])
        val = cell_str(row[1]) if len(row) > 1 else ""
        if label in key_map and val:
            meta[key_map[label]] = val
    return meta


def find_header_row(ws, aliases: dict[str, str]) -> tuple[int, dict[str, int]] | None:
    """返回 (1-based row, {field: col_index 0-based})。aliases: {表头标签: field}。"""
    for r_idx, row in enumerate(ws.iter_rows(min_row=1, max_row=30, max_col=12, values_only=True), 1):
        if not row:
            continue
        colmap: dict[str, int] = {}
        for c_idx, cell in enumerate(row):
            key = aliases.get(cell_str(cell))
            if key:
                colmap[key] = c_idx
        if "set" in colmap and "number" in colmap:
            return r_idx, colmap
    return None


def parse_workbook(path: Path, parse_sheet_fn, merge_fn):
    """遍历工作表。parse_sheet_fn(ws, name) -> (meta, items, errors)；merge_fn(items) -> merged。
    返回 {name: (meta, merged_items)}, errors。"""
    wb = load_workbook(path, read_only=True, data_only=True)
    result: dict[str, tuple[dict, list]] = {}
    errors: list[str] = []
    try:
        for name in wb.sheetnames:
            if should_skip_sheet(name):
                continue
            ws = wb[name]
            meta, items, errs = parse_sheet_fn(ws, name)
            errors.extend(errs)
            if errs:
                continue
            result[name] = (meta, merge_fn(items))
    finally:
        wb.close()
    return result, errors


def write_sheets(sheets, to_txt_fn, person_field, out_dir, default_slug, root):
    """写 txt 文件 + 文件名冲突检测。person_field: 'seller'/'buyer'。"""
    out_dir.mkdir(parents=True, exist_ok=True)
    used: set[str] = set()
    for name, (meta, items) in sheets.items():
        person = meta.get(person_field) or name
        fname = slugify(person, slugify(name, default_slug)) + ".txt"
        # 工作表名若是英文 id 优先用表名
        if re.fullmatch(r"[A-Za-z0-9_-]+", name.strip()):
            fname = name.strip() + ".txt"
        # 冲突检测：同名 slugify 后相同会静默覆盖丢数据，加后缀避免
        if fname in used:
            stem, _, ext = fname.rpartition(".")
            i = 2
            while f"{stem}_{i}.{ext}" in used:
                i += 1
            fname = f"{stem}_{i}.{ext}"
            print(f"  ⚠ 工作表「{name}」文件名冲突，改用 {fname}", file=sys.stderr)
        used.add(fname)
        path = out_dir / fname
        path.write_text(to_txt_fn(meta, items), encoding="utf-8")
        print(f"  写入 {path.relative_to(root) if path.is_relative_to(root) else path}")
