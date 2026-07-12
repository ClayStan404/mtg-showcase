#!/usr/bin/env python3
"""解析 WPS 求购协作 Excel → wants/*.txt（供 build_wants 使用）。

约定（与 templates/WPS求购协作模板.xlsx 一致）：
  - 跳过工作表：说明、模板、名称以「模板」开头
  - 昵称/城市/联系：A4/A5/A6 标签，B 列值
  - 卡表：表头行含「系列」「编号」，其下为数据
  - 列：系列 编号 语言 闪 必须 数量 备注
  - 语言：e=英 z=中 j=日 o=其他（空默认 e）
  - 闪：空/0=否 1=是
  - 必须：空/0=可替其他版 1=必须此版
  - 数量：空=1
"""

from __future__ import annotations

import argparse
import re
import sys
from collections import OrderedDict
from pathlib import Path
from typing import Any

sys.path.insert(0, str(Path(__file__).resolve().parent))

from inventory_format import (  # noqa: E402
    LANG_TOKEN,
    ParseError,
    cell_str,
    normalize_foil,
    normalize_lang,
    normalize_qty,
    normalize_strict,
    slugify,
    validate_meta,
)

try:
    from openpyxl import load_workbook
except ImportError:
    print("需要 openpyxl：pip install openpyxl", file=sys.stderr)
    sys.exit(1)

ROOT = Path(__file__).resolve().parents[1]
SKIP_SHEET_NAMES = {"说明", "使用说明", "填写说明", "对照表（一般不用改）"}
SKIP_SHEET_PREFIXES = ("模板", "template", "_")


def should_skip_sheet(name: str) -> bool:
    n = (name or "").strip()
    if not n or n in SKIP_SHEET_NAMES:
        return True
    low = n.lower()
    for p in SKIP_SHEET_PREFIXES:
        if n.startswith(p) or low.startswith(p):
            return True
    return False


def find_meta(ws) -> dict[str, str]:
    """从工作表前 15 行找 买家昵称/城市/联系。"""
    meta = {"buyer": "", "city": "", "contact": ""}
    key_map = {
        "昵称": "buyer",
        "买家昵称": "buyer",
        "买家": "buyer",
        "buyer": "buyer",
        "城市": "city",
        "city": "city",
        "联系": "contact",
        "联系方式": "contact",
        "contact": "contact",
        "微信": "contact",
    }
    for row in ws.iter_rows(min_row=1, max_row=15, max_col=4, values_only=True):
        if not row:
            continue
        label = cell_str(row[0])
        val = cell_str(row[1]) if len(row) > 1 else ""
        if label in key_map and val:
            meta[key_map[label]] = val
    return meta


def find_header_row(ws) -> tuple[int, dict[str, int]] | None:
    """返回 (1-based row, {field: col_index 0-based})。"""
    aliases = {
        "系列": "set",
        "系列缩写": "set",
        "set": "set",
        "编号": "number",
        "收集编号": "number",
        "number": "number",
        "语言": "lang",
        "lang": "lang",
        "闪": "foil",
        "是否闪卡": "foil",
        "闪卡": "foil",
        "foil": "foil",
        "必须": "must",
        "必须此版": "must",
        "must": "must",
        "数量": "qty",
        "qty": "qty",
        "备注": "note",
        "备注（可选）": "note",
        "note": "note",
    }
    for r_idx, row in enumerate(ws.iter_rows(min_row=1, max_row=30, max_col=12, values_only=True), 1):
        if not row:
            continue
        colmap: dict[str, int] = {}
        for c_idx, cell in enumerate(row):
            key = aliases.get(cell_str(cell))
            if key:
                colmap[key] = c_idx
        if "set" in colmap and "number" in colmap:
            return r_idx, colmap
    return None


def parse_sheet(ws, sheet_name: str) -> tuple[dict[str, str], list[dict[str, Any]], list[str]]:
    """返回 (meta, wants, errors)。"""
    errors: list[str] = []
    meta = find_meta(ws)
    header = find_header_row(ws)
    if not header:
        errors.append(f"[{sheet_name}] 未找到表头（需要「系列」「编号」列）")
        return meta, [], errors

    header_row, colmap = header
    wants: list[dict[str, Any]] = []

    for r_idx, row in enumerate(
        ws.iter_rows(min_row=header_row + 1, max_col=12, values_only=True),
        header_row + 1,
    ):
        if not row or all(cell_str(c) == "" for c in row):
            continue
        try:
            set_code = cell_str(row[colmap["set"]]).lower()
            number = cell_str(row[colmap["number"]])
            if not set_code and not number:
                continue
            if not set_code or not number:
                raise ParseError("系列与编号须同时填写", r_idx)

            lang_raw = row[colmap["lang"]] if "lang" in colmap else ""
            foil_raw = row[colmap["foil"]] if "foil" in colmap else ""
            must_raw = row[colmap["must"]] if "must" in colmap else ""
            qty_raw = row[colmap["qty"]] if "qty" in colmap else ""
            note_raw = row[colmap["note"]] if "note" in colmap else ""

            lang = normalize_lang(lang_raw)
            foil = normalize_foil(foil_raw)
            must = normalize_strict(must_raw)
            qty = normalize_qty(qty_raw)
            note = cell_str(note_raw)

            wants.append(
                {
                    "set": set_code,
                    "number": number,
                    "lang": lang,
                    "foil": foil,
                    "must": must,
                    "quantity": qty,
                    "note": note,
                    "line": r_idx,
                }
            )
        except ParseError as e:
            errors.append(f"[{sheet_name}] 第{r_idx}行：{e}")

    errors.extend(validate_meta(meta, sheet_name, required=("buyer", "city", "contact")))
    return meta, wants, errors


def merge_wants(wants: list[dict[str, Any]]) -> list[dict[str, Any]]:
    merged: OrderedDict[str, dict[str, Any]] = OrderedDict()
    for w in wants:
        key = f"{w['set']}|{w['number']}|{w['lang']}|{'f' if w['foil'] else 'nf'}|{'m' if w['must'] else 'nm'}"
        if key in merged:
            merged[key]["quantity"] += w["quantity"]
        else:
            merged[key] = dict(w)
    return list(merged.values())


def wants_to_txt(meta: dict[str, str], wants: list[dict[str, Any]]) -> str:
    lines = [
        f"# buyer: {meta.get('buyer') or ''}",
        f"# city: {meta.get('city') or ''}",
        f"# contact: {meta.get('contact') or ''}",
        "#",
        "# 由 parse_wps_wants_excel.py 生成 — 语言 e/z/j/o  闪 0/1  必须 0/1  数量默认1",
        "",
    ]
    for w in wants:
        lang_token = LANG_TOKEN.get(w["lang"], "e")
        foil_token = "1" if w["foil"] else "0"
        must_token = "1" if w["must"] else "0"
        qty = w["quantity"]
        prefix = f"{qty}x " if qty != 1 else ""
        # 清洗 note 中的换行（WPS 单元格 Alt+Enter 会被读为 \n），避免断行破坏 txt 格式
        note_raw = (w.get("note") or "").replace("\r", " ").replace("\n", " ").strip()
        note = f" | {note_raw}" if note_raw else ""
        line = f"{prefix}{w['set']} {w['number']} {lang_token} {foil_token} {must_token}{note}"
        lines.append(line)
    lines.append("")
    return "\n".join(lines)


def parse_workbook(path: Path) -> tuple[dict[str, tuple[dict, list]], list[str]]:
    wb = load_workbook(path, read_only=True, data_only=True)
    result: dict[str, tuple[dict, list]] = {}
    errors: list[str] = []
    try:
        for name in wb.sheetnames:
            if should_skip_sheet(name):
                continue
            ws = wb[name]
            meta, wants, errs = parse_sheet(ws, name)
            errors.extend(errs)
            if errs:
                continue
            result[name] = (meta, merge_wants(wants))
    finally:
        wb.close()
    return result, errors


def main() -> int:
    parser = argparse.ArgumentParser(description="WPS 求购 Excel → wants/*.txt")
    parser.add_argument("xlsx", type=Path, help="WPS 导出的求购 xlsx 路径")
    parser.add_argument(
        "-o", "--out-dir", type=Path, default=ROOT / "wants",
        help="输出 wants 目录（默认 ./wants）",
    )
    parser.add_argument("--dry-run", action="store_true", help="只校验不写文件")
    args = parser.parse_args()

    if not args.xlsx.is_file():
        print(f"文件不存在: {args.xlsx}", file=sys.stderr)
        return 1

    print(f"解析 {args.xlsx} …")
    sheets, errors = parse_workbook(args.xlsx)

    if errors:
        print(f"\n❌ 校验失败（{len(errors)} 个问题）：", file=sys.stderr)
        for e in errors:
            print(f"  · {e}", file=sys.stderr)
        return 1

    if not sheets:
        print("没有可导入的工作表（是否只有「说明/模板」？）", file=sys.stderr)
        return 1

    print("\n✅ 校验通过：")
    for name, (meta, wants) in sheets.items():
        total = sum(w["quantity"] for w in wants)
        print(
            f"  · {name}: {meta.get('buyer')} / {meta.get('city')} "
            f"— {len(wants)} 种 / {total} 张"
        )

    if args.dry_run:
        print("\n(--dry-run，未写入文件)")
        return 0

    args.out_dir.mkdir(parents=True, exist_ok=True)
    for name, (meta, wants) in sheets.items():
        buyer = meta.get("buyer") or name
        fname = slugify(buyer, slugify(name, "buyer")) + ".txt"
        if re.fullmatch(r"[A-Za-z0-9_-]+", name.strip()):
            fname = name.strip() + ".txt"
        path = args.out_dir / fname
        path.write_text(wants_to_txt(meta, wants), encoding="utf-8")
        print(f"  写入 {path.relative_to(ROOT) if path.is_relative_to(ROOT) else path}")

    print("\n下一步: python3 scripts/build_wants.py")
    return 0


if __name__ == "__main__":
    sys.exit(main())
